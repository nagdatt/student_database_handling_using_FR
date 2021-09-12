import openpyxl
import datetime
import string
from calendar import monthrange
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill

def autoSize(sheet):
	
	str1=string.ascii_uppercase
	str2=string.ascii_uppercase
	for row in sheet.iter_cols(1): 
		max=""
		isc=0
		for cell in row:
			if(cell.value!=None):
				if(len(cell.value)>=len(max)):
					max=cell.value
				
			print(str1[isc],end=" ")
		
			#my_sheet.column_dimensions[str1[isc]]=len(max)
			isc+=1
		
		print()
		
			
print(sys.argv)

my_wb = openpyxl.load_workbook(sys.argv[4])
my_sheet = my_wb.active
list=['Name','Roll Number','Enrollment Number']
#c2 = my_sheet.cell(row= 1 , column = 2)
print(my_sheet.cell(row= 1 , column = 1).value)
if(my_sheet.cell(row= 1 , column = 1).value==None):
		for j in range(1,4):
           
				my_sheet.cell(row= 1 , column = j).value = list[j-1]
				my_sheet.cell(row= 1 , column = j).font=Font(bold=True,size=12)
		j=1
	
		while(True):
			if(my_sheet.cell(row= j , column = 1).value==None):
						for k in range(1,4):
								my_sheet.cell(row= j , column = k).value = sys.argv[k]
								
						days=1
						for k in range(4, monthrange(int(sys.argv[6]), int(sys.argv[5]))[1]+4):
							date=datetime.datetime.now()
							my_sheet.cell(row= 1 , column = k).value = sys.argv[6]+'/'+sys.argv[5]+'/'+str(days)
							my_sheet.cell(row= 1 , column = k).font=Font(bold=True,size=12)
							days+=1
						break
			else:
				
				
				
				
				
				j+=1
else:
	j=2
	
	while(True):
			if(my_sheet.cell(row= j , column = 1).value==None):
						for k in range(1,4):
								print(sys.argv[k])
								my_sheet.cell(row= j , column = k).value = sys.argv[k]
						break
			else:
				j+=1
l1=['A','B','C','D']
for k in range(1,4):
					my_sheet.column_dimensions[l1[k]].width = 20
autoSize(my_sheet)
my_wb.save(sys.argv[4])



	
